"""Batch operations: re-value, override, and export across multiple companies."""
from __future__ import annotations

import io
from datetime import date as date_type
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from db.session import get_db
from db.models import Company, Valuation
from services.valuation_service import run_company_valuation, apply_override
from services.export_service import export_json, _collect_justifications

router = APIRouter(prefix="/api/v1/batch", tags=["batch"])


# ---------------------------------------------------------------------------
# Request schemas
# ---------------------------------------------------------------------------

class BatchRevalueRequest(BaseModel):
    company_ids: list[UUID]
    created_by: str
    valuation_date: date_type | None = None
    overrides: dict[str, float] | None = None


class BatchOverrideRequest(BaseModel):
    """Apply the same fair-value override to multiple valuations."""
    valuation_ids: list[UUID]
    fair_value: str
    justification: str
    created_by: str


# ---------------------------------------------------------------------------
# Re-value selected companies
# ---------------------------------------------------------------------------

@router.post("/revalue")
def batch_revalue(body: BatchRevalueRequest, db: Session = Depends(get_db)):
    """Re-run valuations for selected companies, optionally with shared overrides."""
    results = []
    for cid in body.company_ids:
        company = db.query(Company).filter(Company.id == cid).first()
        if not company:
            results.append({"company_id": str(cid), "status": "error", "error": "Company not found"})
            continue
        try:
            valuation = run_company_valuation(
                db=db,
                company=company,
                created_by=body.created_by,
                valuation_date=body.valuation_date,
                overrides=body.overrides,
            )
            db.commit()
            methods_run = [
                {
                    "method": mr.get("method", "unknown"),
                    "value": mr.get("value"),
                    "value_low": mr.get("value_low"),
                    "value_high": mr.get("value_high"),
                }
                for mr in (valuation.method_results or [])
            ]
            results.append({
                "company_id": str(company.id),
                "company_name": company.name,
                "valuation_id": str(valuation.id),
                "status": "ok",
                "fair_value": str(valuation.fair_value),
                "fair_value_low": str(valuation.fair_value_low),
                "fair_value_high": str(valuation.fair_value_high),
                "primary_method": valuation.primary_method,
                "explanation": valuation.explanation,
                "methods_run": methods_run,
            })
        except Exception as e:
            db.rollback()
            results.append({
                "company_id": str(cid),
                "company_name": company.name,
                "status": "error",
                "error": str(e),
            })

    return {
        "total": len(body.company_ids),
        "succeeded": sum(1 for r in results if r["status"] == "ok"),
        "failed": sum(1 for r in results if r["status"] == "error"),
        "results": results,
    }


# ---------------------------------------------------------------------------
# Batch override
# ---------------------------------------------------------------------------

@router.post("/override")
def batch_override(body: BatchOverrideRequest, db: Session = Depends(get_db)):
    """Apply the same fair-value override to multiple valuations at once."""
    results = []
    for vid in body.valuation_ids:
        valuation = db.query(Valuation).filter(Valuation.id == vid).first()
        if not valuation:
            results.append({"valuation_id": str(vid), "status": "error", "error": "Valuation not found"})
            continue
        try:
            updated = apply_override(
                db=db,
                valuation=valuation,
                fair_value=body.fair_value,
                justification=body.justification,
                created_by=body.created_by,
            )
            db.commit()
            results.append({
                "valuation_id": str(updated.id),
                "company_name": db.query(Company).filter(Company.id == updated.company_id).first().name,
                "status": "ok",
                "fair_value": str(updated.fair_value),
            })
        except Exception as e:
            db.rollback()
            results.append({"valuation_id": str(vid), "status": "error", "error": str(e)})

    return {
        "total": len(body.valuation_ids),
        "succeeded": sum(1 for r in results if r["status"] == "ok"),
        "failed": sum(1 for r in results if r["status"] == "error"),
        "results": results,
    }


# ---------------------------------------------------------------------------
# Batch export (Excel)
# ---------------------------------------------------------------------------

@router.get("/export")
def batch_export(
    company_ids: str = Query(None, description="Comma-separated company UUIDs. Omit for all."),
    db: Session = Depends(get_db),
):
    """Export latest valuations for selected (or all) companies as a single Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    if company_ids:
        ids = [UUID(x.strip()) for x in company_ids.split(",") if x.strip()]
        companies = db.query(Company).filter(Company.id.in_(ids)).all()
    else:
        companies = db.query(Company).order_by(Company.name).all()

    if not companies:
        raise HTTPException(status_code=404, detail="No companies found")

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    header_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    header_font = Font(bold=True, size=10, color='475569')
    currency_fmt = '#,##0'

    def style_header(ws, row=1, cols=10):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

    # --- Sheet 1: Portfolio Summary ---
    ws = wb.active
    ws.title = "Portfolio Summary"
    headers = ["Company", "Stage", "Sector", "Revenue Status", "Current Revenue",
               "Fair Value", "Low", "High", "Method", "Version", "Valued By", "Valued At"]
    ws.append(headers)
    style_header(ws, row=1, cols=len(headers))

    total_fv = 0
    company_count = 0
    for company in companies:
        latest = (
            db.query(Valuation)
            .filter(Valuation.company_id == company.id)
            .order_by(Valuation.version.desc())
            .first()
        )
        row = [
            company.name,
            company.stage.replace("_", " ").title(),
            company.sector.replace("_", " ").title(),
            company.revenue_status.replace("_", " ").title(),
            float(company.current_revenue) if company.current_revenue else None,
        ]
        if latest:
            row.extend([
                float(latest.fair_value),
                float(latest.fair_value_low),
                float(latest.fair_value_high),
                latest.primary_method.replace("_", " ").title(),
                latest.version,
                latest.created_by,
                latest.created_at.strftime("%Y-%m-%d %H:%M"),
            ])
            total_fv += float(latest.fair_value)
            company_count += 1
        else:
            row.extend([None, None, None, "Not valued", None, None, None])
        ws.append(row)

    # Totals row
    ws.append([])
    ws.append(["TOTAL PORTFOLIO", "", "", "", "", total_fv, "", "", "", f"{company_count} valued"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True, size=10)
    ws.cell(row=ws.max_row, column=6).number_format = currency_fmt

    # Format currency columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, min_col=5, max_col=8):
        for cell in row:
            if cell.value is not None:
                cell.number_format = currency_fmt

    for col, width in [("A", 28), ("B", 16), ("C", 24), ("D", 18), ("E", 18),
                       ("F", 18), ("G", 18), ("H", 18), ("I", 20), ("J", 10),
                       ("K", 14), ("L", 18)]:
        ws.column_dimensions[col].width = width

    # --- Sheet 2: Assumptions & Citations ---
    ws2 = wb.create_sheet("Assumptions & Citations")
    ws2.append(["Company", "Method", "Assumption", "Value", "Rationale", "Source", "Overrideable"])
    style_header(ws2, row=1, cols=7)

    for company in companies:
        latest = (
            db.query(Valuation)
            .filter(Valuation.company_id == company.id)
            .order_by(Valuation.version.desc())
            .first()
        )
        if not latest:
            continue
        for j in _collect_justifications(latest):
            ws2.append([
                company.name,
                j["method"],
                j["assumption"],
                j["value"],
                j["rationale"],
                j["source"],
                "Yes" if j["overrideable"] else "No",
            ])

    for col, width in [("A", 24), ("B", 20), ("C", 24), ("D", 14),
                       ("E", 50), ("F", 50), ("G", 12)]:
        ws2.column_dimensions[col].width = width

    # --- Sheet 3: Method Details ---
    ws3 = wb.create_sheet("Method Details")
    ws3.append(["Company", "Method", "Step", "Formula", "Inputs", "Output"])
    style_header(ws3, row=1, cols=6)

    for company in companies:
        latest = (
            db.query(Valuation)
            .filter(Valuation.company_id == company.id)
            .order_by(Valuation.version.desc())
            .first()
        )
        if not latest:
            continue
        for mr in latest.method_results:
            method_name = mr.get("method", "").replace("_", " ").title()
            for step in mr.get("steps", []):
                inputs_str = ", ".join(f"{k}: {v}" for k, v in step.get("inputs", {}).items())
                ws3.append([
                    company.name,
                    method_name,
                    step.get("description", ""),
                    step.get("formula", ""),
                    inputs_str,
                    step.get("output", ""),
                ])

    for col, width in [("A", 24), ("B", 20), ("C", 30), ("D", 30), ("E", 50), ("F", 24)]:
        ws3.column_dimensions[col].width = width

    # --- Sheet 4: Data Sources ---
    ws4 = wb.create_sheet("Data Sources")
    ws4.append(["Company", "Source Name", "Version", "Effective Date"])
    style_header(ws4, row=1, cols=4)
    seen = set()
    for company in companies:
        latest = (
            db.query(Valuation)
            .filter(Valuation.company_id == company.id)
            .order_by(Valuation.version.desc())
            .first()
        )
        if not latest:
            continue
        for mr in latest.method_results:
            for s in mr.get("sources", []):
                key = f"{company.name}|{s.get('name')}|{s.get('version')}"
                if key not in seen:
                    seen.add(key)
                    ws4.append([
                        company.name,
                        s.get("name", ""),
                        s.get("version", ""),
                        s.get("effective_date", ""),
                    ])

    for col, width in [("A", 24), ("B", 44), ("C", 30), ("D", 16)]:
        ws4.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    filename = f"portfolio-valuations-{date_type.today().isoformat()}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
