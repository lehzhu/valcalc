import io
import json
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from db.models import Valuation, Company


def _format_val(val):
    """Format a value for display, handling nested objects."""
    if val is None:
        return ""
    if isinstance(val, dict):
        return json.dumps(val, indent=1)
    return str(val)


def _collect_justifications(valuation: Valuation) -> list[dict]:
    """Extract all assumptions and their justifications from method results."""
    items = []
    for mr in valuation.method_results:
        method = mr.get("method", "unknown")
        for a in mr.get("assumptions", []):
            items.append({
                "method": method.replace("_", " ").title(),
                "assumption": a.get("name", ""),
                "value": a.get("value", ""),
                "rationale": a.get("rationale", ""),
                "source": a.get("source", ""),
                "overrideable": a.get("overrideable", True),
            })
    return items


def export_json(valuation: Valuation, company: Company, include_justification: bool = True) -> dict:
    """Export valuation as a JSON-serializable dict."""
    result = {
        "company": {
            "name": company.name,
            "stage": company.stage,
            "sector": company.sector,
            "revenue_status": company.revenue_status,
            "current_revenue": str(company.current_revenue) if company.current_revenue else None,
        },
        "valuation": {
            "id": str(valuation.id),
            "version": valuation.version,
            "primary_method": valuation.primary_method,
            "fair_value": str(valuation.fair_value),
            "fair_value_low": str(valuation.fair_value_low),
            "fair_value_high": str(valuation.fair_value_high),
            "explanation": valuation.explanation,
            "created_by": valuation.created_by,
            "created_at": valuation.created_at.isoformat(),
        },
    }

    if include_justification:
        result["valuation"]["method_results"] = valuation.method_results
        result["valuation"]["audit_trail"] = valuation.audit_trail
        result["valuation"]["overrides"] = valuation.overrides
        result["justifications"] = _collect_justifications(valuation)
    else:
        # Minimal: just method names and values, no steps/assumptions/trail
        result["valuation"]["methods"] = [
            {
                "method": mr.get("method"),
                "value": mr.get("value"),
                "value_low": mr.get("value_low"),
                "value_high": mr.get("value_high"),
                "is_primary": mr.get("is_primary"),
            }
            for mr in valuation.method_results
        ]

    return result


def export_xlsx(valuation: Valuation, company: Company, include_justification: bool = True) -> bytes:
    """Export valuation as an Excel workbook."""
    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    header_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    header_font = Font(bold=True, size=10, color='475569')

    def style_header(ws, row=1, cols=10):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ("Field", "Value"),
        ("Company", company.name),
        ("Stage", company.stage.replace("_", " ").title()),
        ("Sector", company.sector.replace("_", " ").title()),
        ("Revenue Status", company.revenue_status.replace("_", " ").title()),
        ("Current Revenue", float(company.current_revenue) if company.current_revenue else ""),
        ("", ""),
        ("Fair Value", float(valuation.fair_value)),
        ("Fair Value (Low)", float(valuation.fair_value_low)),
        ("Fair Value (High)", float(valuation.fair_value_high)),
        ("Primary Method", valuation.primary_method.replace("_", " ").title()),
        ("", ""),
        ("Explanation", valuation.explanation),
        ("", ""),
        ("Created By", valuation.created_by),
        ("Created At", valuation.created_at.isoformat()),
        ("Version", valuation.version),
    ]
    for row in summary_rows:
        ws.append(row)
    style_header(ws, row=1, cols=2)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65

    # Sheet 2: Method Details
    # Each method is an independent valuation approach, clearly separated.
    ws2 = wb.create_sheet("Method Details")
    method_header_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    method_header_font = Font(bold=True, size=11, color='3730A3')
    step_headers = ["Step #", "Description", "Formula", "Inputs", "Output"]

    for mr in valuation.method_results:
        method_name = mr.get("method", "").replace("_", " ").title()
        is_primary = mr.get("is_primary", False)
        primary_tag = " (PRIMARY)" if is_primary else ""
        # Method header row
        ws2.append([f"{method_name}{primary_tag}", "", "", "", "Independent method"])
        r = ws2.max_row
        for c in range(1, 6):
            ws2.cell(row=r, column=c).fill = method_header_fill
            ws2.cell(row=r, column=c).font = method_header_font
        # Column headers
        ws2.append(step_headers)
        style_header(ws2, row=ws2.max_row, cols=5)
        # Steps
        for idx, step in enumerate(mr.get("steps", []), 1):
            inputs_str = ", ".join(f"{k}: {v}" for k, v in step.get("inputs", {}).items())
            ws2.append([
                idx,
                step.get("description", ""),
                step.get("formula", ""),
                inputs_str,
                step.get("output", ""),
            ])
        # Blank row between methods
        ws2.append([])

    for col, width in [("A", 10), ("B", 40), ("C", 30), ("D", 50), ("E", 30)]:
        ws2.column_dimensions[col].width = width

    if include_justification:
        # Sheet 3: Justifications
        ws3 = wb.create_sheet("Justifications")
        ws3.append(["Method", "Assumption", "Value", "Rationale", "Source", "Overrideable"])
        style_header(ws3, row=1, cols=6)
        for j in _collect_justifications(valuation):
            ws3.append([
                j["method"],
                j["assumption"],
                j["value"],
                j["rationale"],
                j["source"],
                "Yes" if j["overrideable"] else "No",
            ])
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws3.column_dimensions[col].width = 25 if col in ("D",) else 18

        # Sheet 4: Audit Trail
        ws4 = wb.create_sheet("Audit Trail")
        trail = valuation.audit_trail or {}
        ws4.append(["Field", "Value"])
        style_header(ws4, row=1, cols=2)
        ws4.append(["Method Selection", trail.get("method_selection_rationale", "")])
        ws4.append(["Benchmark Version", trail.get("benchmark_version", "")])
        ws4.append(["Engine Version", trail.get("engine_version", "")])
        ws4.append(["Timestamp", trail.get("timestamp", "")])
        ws4.append(["", ""])
        ws4.append(["Input Snapshot", ""])
        for key, val in trail.get("input_snapshot", {}).items():
            ws4.append([key, _format_val(val)])
        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
