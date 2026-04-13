"""Batch ingestion: parse a multi-company Excel/CSV, create companies, run valuations.

The input format is one row per company with columns for all relevant fields.
Each company gets both Last Round and Comps analysis where data permits.
"""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from db.models import Company
from services.valuation_service import run_company_valuation


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------

_COLUMNS = [
    ("Company Name", "Required. e.g., Acme Corp"),
    ("Stage", "pre_seed / seed / series_a / series_b / series_c_plus / late_pre_ipo"),
    ("Sector", "GICS: information_technology, healthcare, financials, consumer_discretionary, etc."),
    ("Revenue Status", "pre_revenue / early_revenue / growing_revenue / scaled_revenue"),
    ("Current Annual Revenue", "e.g., 5000000 or $5M"),
    ("Last Round Date", "YYYY-MM-DD"),
    ("Pre-Money Valuation", "e.g., 30000000 or $30M"),
    ("Amount Raised", "e.g., 10000000 or $10M"),
    ("Lead Investor", "e.g., Sequoia Capital"),
    # Financials for calibration
    ("Revenue at Last Round", "Revenue at time of last financing"),
    ("Gross Margin", "Decimal 0-1, e.g., 0.72"),
    ("Runway Months", "e.g., 18"),
    # Qualitative
    ("Board Plan Status", "exceeded / met / missed"),
    ("Customer Concentration", "low / moderate / high"),
    ("Regulatory Risk", "low / moderate / high"),
    # Cap table
    ("Security Type", "e.g., Series A Preferred"),
    ("Liquidation Preferences", "e.g., 1x non-participating"),
    ("Option Pool %", "e.g., 15"),
    # External
    ("Index Movement %", "Sector index movement since round, e.g., 12 for +12%"),
]


def generate_batch_template() -> bytes:
    """Generate an Excel template for bulk company import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies"

    # Header row
    for col_idx, (name, _note) in enumerate(_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=name)

    # Notes row
    for col_idx, (_name, note) in enumerate(_COLUMNS, 1):
        ws.cell(row=2, column=col_idx, value=note)

    # Example rows
    examples = [
        ["Acme AI", "series_a", "information_technology", "growing_revenue",
         "5000000", "2025-06-01", "30000000", "10000000", "Sequoia Capital",
         "2500000", "0.72", "18", "exceeded", "low", "low",
         "Series A Preferred", "1x non-participating", "15", "5"],
        ["Beta Health", "seed", "healthcare", "early_revenue",
         "800000", "2025-09-15", "12000000", "4000000", "a16z Bio",
         "", "0.65", "14", "met", "moderate", "high",
         "SAFE", "", "10", "-3"],
        ["Gamma Fintech", "series_b", "financials", "scaled_revenue",
         "15000000", "2024-12-01", "80000000", "25000000", "Ribbit Capital",
         "10000000", "0.85", "24", "exceeded", "low", "moderate",
         "Series B Preferred", "1x participating", "12", "8"],
    ]
    for row_idx, example in enumerate(examples, 3):
        for col_idx, val in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Format
    for col_idx in range(1, len(_COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

_COL_MAP = {
    "company name": "name",
    "name": "name",
    "stage": "stage",
    "sector": "sector",
    "industry": "sector",
    "revenue status": "revenue_status",
    "revenue_status": "revenue_status",
    "current annual revenue": "current_revenue",
    "current revenue": "current_revenue",
    "annual revenue": "current_revenue",
    "revenue": "current_revenue",
    "last round date": "lr_date",
    "round date": "lr_date",
    "pre-money valuation": "lr_valuation",
    "pre money valuation": "lr_valuation",
    "pre-money": "lr_valuation",
    "amount raised": "lr_amount",
    "round size": "lr_amount",
    "lead investor": "lr_investor",
    "investor": "lr_investor",
    "revenue at last round": "fin_revenue_at_last_round",
    "gross margin": "fin_gross_margin",
    "runway months": "fin_runway_months",
    "runway": "fin_runway_months",
    "board plan status": "qual_board_plan_status",
    "customer concentration": "qual_customer_concentration",
    "regulatory risk": "qual_regulatory_risk",
    "security type": "ct_security_type",
    "liquidation preferences": "ct_liquidation_preferences",
    "option pool %": "ct_option_pool_pct",
    "option pool": "ct_option_pool_pct",
    "index movement %": "ext_index_movement_pct",
    "index movement": "ext_index_movement_pct",
}


def parse_batch_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    """Parse a multi-company file. Returns a list of company dicts."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        return _parse_batch_excel(content)
    elif ext == "csv":
        return _parse_batch_csv(content)
    else:
        raise ValueError(f"Unsupported file type: .{ext}. Upload .xlsx or .csv.")


def _parse_batch_excel(content: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        raise ValueError("File must have a header row and at least one data row")
    return _parse_rows(rows)


def _parse_batch_csv(content: bytes) -> list[dict[str, Any]]:
    import csv
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = [tuple(r) for r in reader]
    if len(rows) < 2:
        raise ValueError("CSV must have a header row and at least one data row")
    return _parse_rows(rows)


def _parse_rows(rows: list[tuple]) -> list[dict[str, Any]]:
    """Parse tabular rows into a list of company dicts."""
    headers = [str(c).strip().lower() if c else "" for c in rows[0]]

    # Map column indices to field names
    col_fields: dict[int, str] = {}
    for idx, h in enumerate(headers):
        field = _COL_MAP.get(h)
        if not field:
            for pattern, mapped in _COL_MAP.items():
                if pattern in h:
                    field = mapped
                    break
        if field:
            col_fields[idx] = field

    if not col_fields:
        raise ValueError("No recognized column headers. Use the batch template.")

    companies: list[dict[str, Any]] = []
    # Skip notes row if it looks like notes (first cell starts with description-like text)
    start = 1
    if len(rows) > 2:
        first_data = str(rows[1][0] or "").strip().lower()
        if first_data.startswith("required") or first_data.startswith("e.g") or first_data.startswith("pre_seed"):
            start = 2

    for row in rows[start:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        raw: dict[str, str] = {}
        for idx, field in col_fields.items():
            if idx < len(row) and row[idx] is not None:
                val = str(row[idx]).strip()
                if val:
                    raw[field] = val

        if "name" not in raw:
            continue  # Skip rows without a company name

        company = _build_company_dict(raw)
        companies.append(company)

    return companies


def _build_company_dict(raw: dict[str, str]) -> dict[str, Any]:
    """Build a structured company dict from flat field values."""
    result: dict[str, Any] = {}

    result["name"] = raw["name"]
    if "stage" in raw:
        result["stage"] = _normalize_stage(raw["stage"])
    if "sector" in raw:
        result["sector"] = _normalize_sector(raw["sector"])
    if "revenue_status" in raw:
        result["revenue_status"] = _normalize_revenue_status(raw["revenue_status"])
    if "current_revenue" in raw:
        num = _parse_money(raw["current_revenue"])
        if num is not None and num > 0:
            result["current_revenue"] = num

    # Last round
    lr: dict[str, Any] = {}
    if "lr_date" in raw:
        lr["date"] = raw["lr_date"]
    if "lr_valuation" in raw:
        num = _parse_money(raw["lr_valuation"])
        if num is not None:
            lr["pre_money_valuation"] = num
    if "lr_amount" in raw:
        num = _parse_money(raw["lr_amount"])
        if num is not None:
            lr["amount_raised"] = num
    if "lr_investor" in raw:
        lr["lead_investor"] = raw["lr_investor"]
    if lr.get("date") and lr.get("pre_money_valuation"):
        lr.setdefault("amount_raised", Decimal("0"))
        result["last_round"] = lr

    # Financials
    financials: dict[str, Any] = {}
    for key in ("fin_revenue_at_last_round", "fin_gross_margin", "fin_runway_months"):
        if key in raw:
            fin_key = key[4:]
            num = _parse_money(raw[key])
            financials[fin_key] = str(num) if num is not None else raw[key]
    if financials:
        result["financials"] = financials

    # Qualitative
    qualitative: dict[str, str] = {}
    for key in ("qual_board_plan_status", "qual_customer_concentration", "qual_regulatory_risk"):
        if key in raw:
            qualitative[key[5:]] = raw[key]
    if qualitative:
        result["qualitative"] = qualitative

    # Cap table
    cap_table: dict[str, str] = {}
    for key in ("ct_security_type", "ct_liquidation_preferences", "ct_option_pool_pct"):
        if key in raw:
            cap_table[key[3:]] = raw[key]
    if cap_table:
        result["cap_table"] = cap_table

    # External
    external: dict[str, Any] = {}
    if "ext_index_movement_pct" in raw:
        num = _parse_money(raw["ext_index_movement_pct"])
        if num is not None:
            # If > 1, treat as percentage points
            if abs(num) > 1:
                num = num / 100
            external["index_movement_pct"] = str(num)
    if external:
        result["external_mapping"] = external

    return result


# ---------------------------------------------------------------------------
# Batch execution
# ---------------------------------------------------------------------------

def run_batch_valuation(
    db: Session,
    companies_data: list[dict[str, Any]],
    created_by: str,
    valuation_date: date | None = None,
) -> list[dict[str, Any]]:
    """Create companies and run valuations for each. Returns summary results."""
    results: list[dict[str, Any]] = []

    for data in companies_data:
        try:
            company = _create_or_update_company(db, data, created_by)
            valuation = run_company_valuation(
                db=db,
                company=company,
                created_by=created_by,
                valuation_date=valuation_date,
            )

            # Build per-method breakdown
            methods_run = []
            for mr in (valuation.method_results or []):
                methods_run.append({
                    "method": mr.get("method", mr.get("method_type", "unknown")),
                    "value": mr.get("value"),
                    "value_low": mr.get("value_low"),
                    "value_high": mr.get("value_high"),
                })

            results.append({
                "company_id": str(company.id),
                "company_name": company.name,
                "status": "ok",
                "fair_value": str(valuation.fair_value),
                "fair_value_low": str(valuation.fair_value_low),
                "fair_value_high": str(valuation.fair_value_high),
                "primary_method": valuation.primary_method,
                "explanation": valuation.explanation,
                "methods_run": methods_run,
                "valuation_id": str(valuation.id),
            })
        except Exception as e:
            results.append({
                "company_name": data.get("name", "Unknown"),
                "status": "error",
                "error": str(e),
            })

    return results


def _create_or_update_company(
    db: Session,
    data: dict[str, Any],
    created_by: str,
) -> Company:
    """Create a new company or update an existing one by name."""
    existing = db.query(Company).filter(Company.name == data["name"]).first()

    if existing:
        company = existing
    else:
        company = Company(
            name=data["name"],
            stage=data.get("stage", "seed"),
            sector=data.get("sector", "information_technology"),
            revenue_status=data.get("revenue_status", "pre_revenue"),
            created_by=created_by,
        )
        db.add(company)

    # Update fields
    if "stage" in data:
        company.stage = data["stage"]
    if "sector" in data:
        company.sector = data["sector"]
    if "revenue_status" in data:
        company.revenue_status = data["revenue_status"]
    if "current_revenue" in data:
        company.current_revenue = data["current_revenue"]

    if "last_round" in data:
        lr = data["last_round"]
        company.last_round_date = lr["date"] if isinstance(lr["date"], date) else date.fromisoformat(lr["date"])
        company.last_round_valuation = lr["pre_money_valuation"]
        company.last_round_amount = lr.get("amount_raised", Decimal("0"))
        company.last_round_investor = lr.get("lead_investor")

    if "financials" in data:
        company.financials = data["financials"]
    if "qualitative" in data:
        company.qualitative = data["qualitative"]
    if "cap_table" in data:
        company.cap_table = data["cap_table"]
    if "external_mapping" in data:
        company.external_mapping = data["external_mapping"]

    db.flush()
    return company


# ---------------------------------------------------------------------------
# Helpers (shared with document_parser)
# ---------------------------------------------------------------------------

def _parse_money(s: str) -> Decimal | None:
    s = s.strip().replace(",", "").replace("$", "")
    multiplier = 1
    if s.upper().endswith("B"):
        multiplier = 1_000_000_000
        s = s[:-1]
    elif s.upper().endswith("M"):
        multiplier = 1_000_000
        s = s[:-1]
    elif s.upper().endswith("K"):
        multiplier = 1_000
        s = s[:-1]
    try:
        return Decimal(s.strip()) * multiplier
    except (InvalidOperation, ValueError):
        return None


_STAGE_ALIASES = {
    "pre-seed": "pre_seed", "preseed": "pre_seed", "pre seed": "pre_seed",
    "seed": "seed",
    "series a": "series_a", "a": "series_a",
    "series b": "series_b", "b": "series_b",
    "series c": "series_c_plus", "series c+": "series_c_plus", "c+": "series_c_plus",
    "series d": "series_c_plus", "series e": "series_c_plus",
    "late": "late_pre_ipo", "pre-ipo": "late_pre_ipo", "late / pre-ipo": "late_pre_ipo",
}

def _normalize_stage(s: str) -> str:
    s_lower = s.lower().strip()
    return _STAGE_ALIASES.get(s_lower, s_lower.replace(" ", "_"))

_REVENUE_ALIASES = {
    "pre-revenue": "pre_revenue", "none": "pre_revenue",
    "early": "early_revenue", "<$1m": "early_revenue",
    "growing": "growing_revenue", "$1-10m": "growing_revenue",
    "scaled": "scaled_revenue", ">$10m": "scaled_revenue",
}

def _normalize_revenue_status(s: str) -> str:
    s_lower = s.lower().strip()
    return _REVENUE_ALIASES.get(s_lower, s_lower.replace(" ", "_"))

def _normalize_sector(s: str) -> str:
    s_lower = s.lower().strip().replace(" ", "_").replace("-", "_")
    aliases = {
        "tech": "information_technology", "software": "information_technology",
        "saas": "information_technology", "it": "information_technology",
        "ai": "information_technology", "technology": "information_technology",
        "health": "healthcare", "biotech": "healthcare", "pharma": "healthcare",
        "finance": "financials", "fintech": "financials",
        "consumer": "consumer_discretionary", "retail": "consumer_discretionary",
        "ecommerce": "consumer_discretionary",
        "media": "communication_services", "telecom": "communication_services",
        "clean_energy": "energy", "cleantech": "energy", "climate": "energy",
        "hardware": "industrials", "manufacturing": "industrials",
    }
    return aliases.get(s_lower, s_lower)
