"""Parse uploaded Excel / CSV files into structured company data.

Supports two document layouts:
  1. **Projections sheet** – columns: Year, Revenue, EBITDA, Growth Rate
  2. **Company summary sheet** – key-value pairs for round info, revenue, stage, sector

The parser auto-detects layout by scanning header rows.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_upload(filename: str, content: bytes) -> dict[str, Any]:
    """Parse an uploaded file and return extracted company data.

    Returns a dict with optional keys matching CompanyCreate / CompanyUpdate:
      - name, stage, sector, revenue_status
      - current_revenue
      - last_round: {date, pre_money_valuation, amount_raised, lead_investor}
      - projections: {periods: [{year, revenue, ebitda, growth_rate}, ...]}
    Only keys that were successfully extracted are included.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        return _parse_excel(content)
    elif ext == "csv":
        return _parse_csv(content)
    else:
        raise ValueError(f"Unsupported file type: .{ext}. Upload .xlsx or .csv files.")


def generate_template() -> bytes:
    """Generate a comprehensive Excel template with all 5 data categories."""
    wb = openpyxl.Workbook()

    # Sheet 1 – Transaction / Round Data
    ws1 = wb.active
    ws1.title = "Transaction"
    for row in [
        ("Field", "Value", "Notes"),
        ("Company Name", "", "e.g., Acme Corp"),
        ("Stage", "", "pre_seed / seed / series_a / series_b / series_c_plus / late_pre_ipo"),
        ("Sector", "", "GICS: information_technology, healthcare, financials, consumer_discretionary, etc."),
        ("Revenue Status", "", "pre_revenue / early_revenue / growing_revenue / scaled_revenue"),
        ("Last Round Date", "", "YYYY-MM-DD"),
        ("Last Round Pre-Money Valuation", "", "e.g., $30M or 30000000"),
        ("Last Round Amount Raised", "", "e.g., $10M or 10000000"),
        ("Last Round Lead Investor", "", "e.g., Sequoia Capital"),
        ("Security Type", "", "e.g., Series A Preferred"),
        ("Liquidation Preferences", "", "e.g., 1x non-participating"),
        ("Option Pool %", "", "e.g., 15"),
        ("SAFEs / Convertible Notes", "", "e.g., $2M SAFE at $8M cap"),
        ("Convertibility", "", "e.g., Mandatory at IPO"),
    ]:
        ws1.append(row)
    for col in range(1, 4):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 34

    # Sheet 2 – Financials
    ws2 = wb.create_sheet("Financials")
    for row in [
        ("Field", "Value", "Notes"),
        ("Current Annual Revenue", "", "e.g., $5M or 5000000"),
        ("Revenue at Last Round", "", "Revenue at the time of last financing"),
        ("Gross Margin", "", "Decimal 0-1, e.g., 0.72"),
        ("Net Burn Rate (monthly)", "", "e.g., $200K or 200000"),
        ("Runway (months)", "", "e.g., 18"),
        ("Cash on Hand", "", "e.g., $3.6M"),
        ("EBITDA (trailing)", "", "Can be negative"),
        ("ARR", "", "If different from revenue"),
        ("MRR", "", "Monthly recurring revenue"),
        ("Revenue Growth YoY %", "", "e.g., 120 for 120%"),
    ]:
        ws2.append(row)
    for col in range(1, 4):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 34

    # Sheet 3 – Forecast / Projections
    ws3 = wb.create_sheet("Forecast")
    ws3.append(("Year", "Revenue", "EBITDA", "Growth Rate (%)"))
    current_year = date.today().year
    for y in range(current_year + 1, current_year + 6):
        ws3.append((y, "", "", ""))
    for col in range(1, 5):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # Sheet 4 – Qualitative
    ws4 = wb.create_sheet("Qualitative")
    for row in [
        ("Field", "Value", "Notes"),
        ("Board Plan Status", "", "exceeded / met / missed / on track"),
        ("Customer Concentration", "", "low / moderate / high"),
        ("Regulatory Risk", "", "low / moderate / high / material"),
        ("Major Events Since Round", "", "Free text: key milestones, pivots, etc."),
        ("Key Risks", "", "Free text: competitive threats, tech risk, etc."),
        ("Management Changes", "", "Free text: new hires, departures"),
    ]:
        ws4.append(row)
    for col in range(1, 4):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 34

    # Sheet 5 – External Mapping
    ws5 = wb.create_sheet("External Mapping")
    for row in [
        ("Field", "Value", "Notes"),
        ("Market Index", "", "e.g., NASDAQ Composite, S&P 500 IT"),
        ("Index Movement Since Round %", "", "e.g., 12 for +12%"),
        ("Comparable Public Companies", "", "Comma-separated tickers"),
        ("Private Transaction Comps", "", "Recent private deals for context"),
        ("409A Valuation (if available)", "", "Most recent 409A fair value"),
        ("409A Date", "", "YYYY-MM-DD"),
    ]:
        ws5.append(row)
    for col in range(1, 4):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 34

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _parse_excel(content: bytes) -> dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    result: dict[str, Any] = {}

    for ws in wb.worksheets:
        title = ws.title.lower().strip()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect layout from headers
        headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        # Batch/columnar layout (one row per company) — extract first company
        if _is_batch_layout(headers):
            from services.batch_service import parse_batch_file
            companies, _ = parse_batch_file("upload.xlsx", content)
            if companies:
                result.update(companies[0])
            wb.close()
            return result

        if _is_projection_headers(headers):
            projections = _extract_projections_from_rows(headers, rows[1:])
            if projections:
                result["projections"] = {"periods": projections}
        elif _is_kv_layout(headers) or "company" in title or "info" in title or "summary" in title:
            kv = _extract_kv_from_rows(rows)
            result.update(kv)
        else:
            # Try projections first, fall back to kv
            projections = _extract_projections_from_rows(headers, rows[1:])
            if projections:
                result["projections"] = {"periods": projections}
            else:
                kv = _extract_kv_from_rows(rows)
                if kv:
                    result.update(kv)

    wb.close()
    return result


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------

def _parse_csv(content: bytes) -> dict[str, Any]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("Empty CSV file")

    headers = [c.strip().lower() for c in rows[0]]

    # Batch/columnar layout — extract first company
    if _is_batch_layout(headers):
        from services.batch_service import parse_batch_file
        companies, _ = parse_batch_file("upload.csv", content)
        if companies:
            return companies[0]
        return {}

    if _is_projection_headers(headers):
        projections = _extract_projections_from_rows(headers, [tuple(r) for r in rows[1:]])
        if projections:
            return {"projections": {"periods": projections}}
    # Fall back to key-value
    kv = _extract_kv_from_rows([tuple(r) for r in rows])
    return kv


# ---------------------------------------------------------------------------
# Layout detection
# ---------------------------------------------------------------------------

_PROJECTION_KEYWORDS = {"year", "revenue", "ebitda", "growth"}

# Column headers that indicate a batch/columnar layout (one row per company)
_BATCH_COLUMN_KEYWORDS = {
    "company name", "stage", "sector", "revenue status",
    "pre-money valuation", "amount raised", "last round date",
}

def _is_projection_headers(headers: list[str]) -> bool:
    return len(set(headers) & _PROJECTION_KEYWORDS) >= 2


def _is_batch_layout(headers: list[str]) -> bool:
    """Detect a columnar/batch layout (multiple recognized column headers)."""
    matches = sum(1 for h in headers if h in _BATCH_COLUMN_KEYWORDS)
    return matches >= 3


def _is_kv_layout(headers: list[str]) -> bool:
    return any(h in ("field", "key", "parameter", "item") for h in headers)


# ---------------------------------------------------------------------------
# Projection extraction
# ---------------------------------------------------------------------------

def _extract_projections_from_rows(
    headers: list[str], data_rows: list[tuple],
) -> list[dict[str, Any]]:
    col_map = {}
    for i, h in enumerate(headers):
        h_clean = h.replace(" ", "_").replace("(%)", "").strip("_")
        if "year" in h_clean:
            col_map["year"] = i
        elif "revenue" in h_clean and "growth" not in h_clean:
            col_map["revenue"] = i
        elif "ebitda" in h_clean:
            col_map["ebitda"] = i
        elif "growth" in h_clean:
            col_map["growth_rate"] = i

    if "year" not in col_map or "revenue" not in col_map:
        return []

    periods = []
    for row in data_rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        try:
            year = int(float(str(row[col_map["year"]])))
        except (ValueError, TypeError, IndexError):
            continue

        rev_raw = _cell_value(row, col_map["revenue"])
        if rev_raw is None:
            continue

        period: dict[str, Any] = {"year": year, "revenue": str(rev_raw)}

        if "ebitda" in col_map:
            ebitda = _cell_value(row, col_map["ebitda"])
            if ebitda is not None:
                period["ebitda"] = str(ebitda)

        if "growth_rate" in col_map:
            gr = _cell_value(row, col_map["growth_rate"])
            if gr is not None:
                # Normalize: if > 1, assume it's a percentage (e.g. 25 -> 0.25)
                if gr > 1:
                    gr = gr / 100
                period["growth_rate"] = float(gr)

        periods.append(period)

    return periods


# ---------------------------------------------------------------------------
# Key-value extraction
# ---------------------------------------------------------------------------

_FIELD_MAP = {
    # Core company info
    "company name": "name",
    "name": "name",
    "stage": "stage",
    "sector": "sector",
    "industry": "sector",
    "revenue status": "revenue_status",
    "revenue_status": "revenue_status",
    # Last round
    "last round date": "lr_date",
    "round date": "lr_date",
    "funding date": "lr_date",
    "last round pre-money valuation": "lr_valuation",
    "last round valuation": "lr_valuation",
    "pre-money valuation": "lr_valuation",
    "pre money valuation": "lr_valuation",
    "pre-money": "lr_valuation",
    "last round amount raised": "lr_amount",
    "amount raised": "lr_amount",
    "round size": "lr_amount",
    "last round lead investor": "lr_investor",
    "lead investor": "lr_investor",
    "investor": "lr_investor",
    # Cap table (→ cap_table JSON blob)
    "security type": "ct_security_type",
    "liquidation preferences": "ct_liquidation_preferences",
    "option pool %": "ct_option_pool_pct",
    "option pool": "ct_option_pool_pct",
    "safes / convertible notes": "ct_safes_notes",
    "safes": "ct_safes_notes",
    "convertible notes": "ct_safes_notes",
    "convertibility": "ct_convertibility",
    # Financials (→ financials JSON blob)
    "current revenue": "fin_current_revenue",
    "current annual revenue": "fin_current_revenue",
    "annual revenue": "fin_current_revenue",
    "revenue": "fin_current_revenue",
    "revenue at last round": "fin_revenue_at_last_round",
    "gross margin": "fin_gross_margin",
    "net burn rate": "fin_burn_rate",
    "burn rate": "fin_burn_rate",
    "runway": "fin_runway_months",
    "runway (months)": "fin_runway_months",
    "runway months": "fin_runway_months",
    "cash on hand": "fin_cash_on_hand",
    "ebitda": "fin_ebitda",
    "ebitda (trailing)": "fin_ebitda",
    "arr": "fin_arr",
    "mrr": "fin_mrr",
    "revenue growth yoy %": "fin_revenue_growth_yoy",
    "revenue growth": "fin_revenue_growth_yoy",
    # Qualitative (→ qualitative JSON blob)
    "board plan status": "qual_board_plan_status",
    "customer concentration": "qual_customer_concentration",
    "regulatory risk": "qual_regulatory_risk",
    "major events since round": "qual_major_events",
    "major events": "qual_major_events",
    "key risks": "qual_key_risks",
    "management changes": "qual_management_changes",
    # External mapping (→ external_mapping JSON blob)
    "market index": "ext_index_name",
    "index movement since round %": "ext_index_movement_pct",
    "index movement": "ext_index_movement_pct",
    "comparable public companies": "ext_public_comps",
    "private transaction comps": "ext_private_comps",
    "409a valuation": "ext_409a_value",
    "409a valuation (if available)": "ext_409a_value",
    "409a date": "ext_409a_date",
}


def _extract_kv_from_rows(rows: list[tuple]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    lr: dict[str, Any] = {}
    cap_table: dict[str, Any] = {}
    financials: dict[str, Any] = {}
    qualitative: dict[str, Any] = {}
    external: dict[str, Any] = {}

    for row in rows:
        if len(row) < 2:
            continue
        key = str(row[0]).strip().lower() if row[0] is not None else ""
        val = row[1]
        if not key or val is None or str(val).strip() == "":
            continue

        field = _FIELD_MAP.get(key)
        if not field:
            # Fuzzy: strip extra words
            for pattern, mapped in _FIELD_MAP.items():
                if pattern in key:
                    field = mapped
                    break
        if not field:
            continue

        val_str = str(val).strip()

        # Core fields
        if field == "name":
            result["name"] = val_str
        elif field == "stage":
            result["stage"] = _normalize_stage(val_str)
        elif field == "sector":
            result["sector"] = _normalize_sector(val_str)
        elif field == "revenue_status":
            result["revenue_status"] = _normalize_revenue_status(val_str)

        # Last round
        elif field == "lr_date":
            d = _parse_date(val_str)
            if d:
                lr["date"] = d.isoformat()
        elif field == "lr_valuation":
            num = _parse_money(val_str)
            if num is not None:
                lr["pre_money_valuation"] = str(num)
        elif field == "lr_amount":
            num = _parse_money(val_str)
            if num is not None:
                lr["amount_raised"] = str(num)
        elif field == "lr_investor":
            lr["lead_investor"] = val_str

        # Cap table
        elif field.startswith("ct_"):
            cap_key = field[3:]  # strip "ct_"
            cap_table[cap_key] = val_str

        # Financials
        elif field.startswith("fin_"):
            fin_key = field[4:]  # strip "fin_"
            num = _parse_money(val_str)
            if num is not None:
                financials[fin_key] = str(num)
                # Also set top-level current_revenue
                if fin_key == "current_revenue":
                    result["current_revenue"] = str(num)
            else:
                financials[fin_key] = val_str

        # Qualitative
        elif field.startswith("qual_"):
            qual_key = field[5:]  # strip "qual_"
            qualitative[qual_key] = val_str

        # External mapping
        elif field.startswith("ext_"):
            ext_key = field[4:]  # strip "ext_"
            if ext_key in ("index_movement_pct", "409a_value"):
                num = _parse_money(val_str)
                if num is not None:
                    external[ext_key] = str(num)
                else:
                    external[ext_key] = val_str
            elif ext_key == "409a_date":
                d = _parse_date(val_str)
                if d:
                    external[ext_key] = d.isoformat()
            else:
                external[ext_key] = val_str

    if lr.get("date") and lr.get("pre_money_valuation"):
        lr.setdefault("amount_raised", "0")
        result["last_round"] = lr
    if cap_table:
        result["cap_table"] = cap_table
    if financials:
        result["financials"] = financials
    if qualitative:
        result["qualitative"] = qualitative
    if external:
        result["external_mapping"] = external

    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_value(row: tuple, idx: int) -> Decimal | None:
    if idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("$", "").replace("%", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_money(s: str) -> Decimal | None:
    """Parse money strings like '$30M', '5,000,000', '10000000'."""
    s = s.strip().replace(",", "").replace("$", "")
    # Handle suffixes: K, M, B
    multiplier = 1
    if s.upper().endswith("B"):
        multiplier = 1_000_000_000
        s = s[:-1]
    elif s.upper().endswith("M"):
        multiplier = 1_000_000
        s = s[:-1]
    elif s.upper().endswith("K"):
        multiplier = 1_000
        s = s[:-1]
    try:
        return Decimal(s.strip()) * multiplier
    except (InvalidOperation, ValueError):
        return None


def _parse_date(s: str) -> date | None:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    # Try Excel serial date
    try:
        serial = int(float(s))
        if 30000 < serial < 60000:
            return datetime(1899, 12, 30) + __import__("datetime").timedelta(days=serial)
    except (ValueError, TypeError):
        pass
    return None


_STAGE_ALIASES = {
    "pre-seed": "pre_seed", "preseed": "pre_seed", "pre seed": "pre_seed",
    "seed": "seed",
    "series a": "series_a", "a": "series_a",
    "series b": "series_b", "b": "series_b",
    "series c": "series_c_plus", "series c+": "series_c_plus", "c": "series_c_plus", "c+": "series_c_plus",
    "series d": "series_c_plus", "series e": "series_c_plus",
    "late": "late_pre_ipo", "pre-ipo": "late_pre_ipo", "pre ipo": "late_pre_ipo",
    "late / pre-ipo": "late_pre_ipo", "ipo": "late_pre_ipo",
}

def _normalize_stage(s: str) -> str:
    s_lower = s.lower().strip()
    if s_lower in _STAGE_ALIASES:
        return _STAGE_ALIASES[s_lower]
    # Already a valid key?
    valid = {"pre_seed", "seed", "series_a", "series_b", "series_c_plus", "late_pre_ipo"}
    if s_lower.replace(" ", "_") in valid:
        return s_lower.replace(" ", "_")
    return s_lower


_REVENUE_ALIASES = {
    "pre-revenue": "pre_revenue", "pre revenue": "pre_revenue", "none": "pre_revenue",
    "early": "early_revenue", "early revenue": "early_revenue", "<$1m": "early_revenue",
    "growing": "growing_revenue", "growing revenue": "growing_revenue", "$1-10m": "growing_revenue",
    "scaled": "scaled_revenue", "scaled revenue": "scaled_revenue", ">$10m": "scaled_revenue",
}

def _normalize_revenue_status(s: str) -> str:
    s_lower = s.lower().strip()
    if s_lower in _REVENUE_ALIASES:
        return _REVENUE_ALIASES[s_lower]
    valid = {"pre_revenue", "early_revenue", "growing_revenue", "scaled_revenue"}
    if s_lower.replace(" ", "_") in valid:
        return s_lower.replace(" ", "_")
    return s_lower


def _normalize_sector(s: str) -> str:
    s_lower = s.lower().strip().replace(" ", "_").replace("-", "_").replace("/", "_")
    # Map common names to GICS sector keys
    aliases = {
        "it": "information_technology", "tech": "information_technology",
        "technology": "information_technology", "software": "information_technology",
        "saas": "information_technology", "b2b_saas": "information_technology",
        "ai": "information_technology", "ai_ml": "information_technology",
        "cybersecurity": "information_technology", "enterprise_software": "information_technology",
        "health": "healthcare", "biotech": "healthcare", "pharma": "healthcare",
        "healthtech": "healthcare", "healthcare_biotech": "healthcare",
        "finance": "financials", "financial": "financials", "fintech": "financials",
        "consumer": "consumer_discretionary", "retail": "consumer_discretionary",
        "ecommerce": "consumer_discretionary", "e_commerce": "consumer_discretionary",
        "ecommerce_marketplace": "consumer_discretionary",
        "media": "communication_services", "telecom": "communication_services",
        "consumer_tech": "communication_services",
        "clean_energy": "energy", "cleantech": "energy", "climate": "energy",
        "climate_cleantech": "energy",
        "hardware": "industrials", "manufacturing": "industrials",
        "hardware_iot": "industrials", "iot": "industrials",
    }
    return aliases.get(s_lower, s_lower)
